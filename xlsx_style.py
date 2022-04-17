import os, sys
import json
import time
import re
import openpyxl
from openpyxl.styles import NamedStyle, Font, PatternFill, Alignment, Side, Border

##------------------------------------------------------------
## User 모듈
##------------------------------------------------------------
sys.path.append(os.path.join(os.path.dirname(__file__), '.')) ## Note: 현재 디렉토리 기준 상대 경로 설정
from utils_basic import (
    _create_folder, 
    _read_file, 
    _file_to_json, 
    _to_lists,
    _to_dicts,
    _to_df, 
    _project_dicts,
    _fn
)

##@@@ 전역 상수/변수
##============================================================

##@@ setting, 
##------------------------------------------------------------

WIDTH_PX_RATE = 0.12
HEIGHT_PX_RATE = 0.75


##@@ theme, 
##------------------------------------------------------------

themes = {
    'default': {
        'header': dict(
            font = dict(color="DAEEF3", bold=True, size=12),
            patternfill = dict(fill_type="solid", fgColor="050A30"),
            alignment = dict(horizontal='center', vertical='center', wrap_text=True),
            sides = [Side(style='thin', color="DAEEF3")],
            border = dict(left='sides[0]', top='sides[0]', right='sides[0]', bottom='sides[0]')
        ),
        'rows': dict(
            font = dict(color="050A30", bold=False, size=12),
            patternfill = dict(fill_type="solid", fgColor="EFF9FF"),
            alignment = dict(horizontal='left', vertical='center', wrap_text=True),
            sides = [Side(style='thin', color="050A30")],
            border = dict(left='sides[0]', top='sides[0]', right='sides[0]', bottom='sides[0]')
        )
    },
    'pink': {
        'header': dict(
            font = dict(color="FFEBFF", bold=True, size=12),
            patternfill = dict(fill_type="solid", fgColor="660066"),
            alignment = dict(horizontal='center', vertical='center', wrap_text=True),
            sides = [Side(style='thin', color="FFEBFF")],
            border = dict(left='sides[0]', top='sides[0]', right='sides[0]', bottom='sides[0]')
        ),
        'rows': dict(
            font = dict(color="660066", bold=False, size=12),
            patternfill = dict(fill_type="solid", fgColor="FFEBFF"),
            alignment = dict(horizontal='left', vertical='center', wrap_text=True),
            sides = [Side(style='thin', color="660066")],
            border = dict(left='sides[0]', top='sides[0]', right='sides[0]', bottom='sides[0]')
        )
    }

}


## openpyxl
##--------------------------------------------------------------
def _xsheet(path=None, title=''):
    """worksheet 오브젝트

    Args:
        path (str, optional): workbook path / None: 새 workbook. Defaults to None.
        title (str, optional): worksheet 이름. Defaults to ''.

    Returns:
        [obj]: worksheet 오브젝트

    Usages:
        sheet = _xsheet('', '')
        ws = sheet['ws']
        wb = sheet['wb']
    """
    # print(f"{_fn(1)} path: {path}")
    if path == None:  # wb, ws 생성
        wb = openpyxl.Workbook()
        ws = wb.worksheets[0]
        ws.title = title
    else:  # wb, ws 로딩
        try:
            wb = openpyxl.load_workbook(filename=path)
            if title in wb.sheetnames:
                ws = wb[title]
                # print(f"{_fn(1)} ws = wb[title]")
            else:
                # print(f"{_fn(1)} create_sheet")
                ws = wb.create_sheet(title) # 마지막
                # wb.create_sheet(title, 0) # insert at first position
                # wb.create_sheet(title, -1) # 끝에서 2번째
        except Exception as e:  ## NOTE: 파일이 없는 경우
            wb = openpyxl.Workbook()
            ws = wb.worksheets[0]
            ws.title = title

    return {'wb': wb, 'ws': ws}


def _copy_ws_style(src, dest):
    # new_sheet = workbook.create_sheet(sheetName)
    # default_sheet = workbook['default']

    for row in src.rows:
        for cell in row:
            new_cell = src.cell(row=cell.row_idx, col=cell.col_idx, value= cell.value)
            if cell.has_style:
                new_cell.font = cell.font
                new_cell.border = cell.border
                new_cell.fill = cell.fill
                new_cell.number_format = cell.number_format
                new_cell.protection = cell.protection
                new_cell.alignment = cell.alignment


def _copy_before_row_style(ws, from_to=(2, 5)):
    """데이터 추가 전 마지막 행의 스타일을 데이터 추가시 사용

    Args:
        ws ([type]): [description]
        from_to (tuple, optional): 추가되는 행 번호(from, to) . Defaults to (2, 5).
    """

    src_row = ws[from_to[0]-1]
    for i in range(from_to[0], from_to[1]+1):
        for j, src_cell in enumerate(src_row):
            ws.cell(row = i, column = j+1).style = src_cell.style


## NOTE: 사용자 스타일 설정
def _set_named_style(wb, name, options):
    """사용자 스타일 설정

    Args:
        wb (obj): workbook 오브젝트
        name (str): 스타일 이름
        options (dict): 스타일 옵션

    Returns:
        [obj]: NamedStyle 오브젝트
    
    Usages:
        # 옵션 설정
        options = dict(
            font = dict(color="DAEEF3", bold=True, size=12),
            patternfill = dict(fill_type="solid", fgColor="050A30"),
            alignment = dict(horizontal='center', vertical='center'),
            sides = [Side(style='thin', color="FF0000")],
            border = dict(left='sides[0]', right='sides[0]')
        )
        # 스타일 추가 실행
        _set_named_style(wb, 'style_header', options)

        # cell에 스타일 적용
        ws["D1"].style = 'style_header'
    """
    _style = NamedStyle(name=name)
    if 'font' in options:
        _style.font = Font(**options['font'])
    if 'patternfill' in options:
        _style.fill = PatternFill(**options['patternfill'])
    if 'alignment' in options:
        _style.alignment = Alignment(**options['alignment'])
    if 'border' in options:
        if 'sides' in options:
            sides = options['sides']
        else:
            sides = [Side(style='thin', color="DAEEF3")]
        
        borders = {}
        for key, val in options['border'].items():
            borders[key] = eval(val)

        _style.border = Border(**borders)

    try:  ## TODO: style 이름이 존재하면 처음부터 처리하지 않도록
        wb.add_named_style(_style)
    except:
        # print("이미 존재하는 사용자 스타일입니다. 업데이트합니다.")
        del wb._named_styles[wb.style_names.index(name)]
        wb.add_named_style(_style)

    return _style


# _apply_style_theme(wb, theme, count, header, first)
def _apply_style_theme(sheet, theme, count_row, count_col, header, first):
    """스타일 테마 적용

    Args:
        wb ([type]): [description]
        theme ([type]): [description]
    """
    wb = sheet['wb']
    ws = sheet['ws']

    _set_named_style(wb, 'header_style', themes[theme]['header'])
    _set_named_style(wb, 'rows_style', themes[theme]['rows'])
    ## header
    if header != None:
        for i in range(first[0], first[0] + header+1):
            for j in range(first[1], count_col):  # TODO: first 열이 0이 아닌 경우
                col_letter = openpyxl.utils.get_column_letter(j + 1)
                ws[f"{col_letter}{i+1}"].style = 'header_style'
    else:  # header는 적용하지 않는 경우
        header = 0
    
    ## rows
    for i in range(first[0] + header+1, first[0] + header+ count_row):
        for j in range(first[1], count_col):  # TODO: first 열이 0이 아닌 경우
            col_letter = openpyxl.utils.get_column_letter(j + 1)
            ws[f"{col_letter}{i+1}"].style = 'rows_style'


def _fit_width(ws):
    """너비 자동맞춤

    Args:
        ws (obj): worksheet 오브젝트
    """
    column_letters = tuple(openpyxl.utils.get_column_letter(col_number + 1) for col_number in range(ws.max_column))
    for column_letter in column_letters:
        ws.column_dimensions[column_letter].bestFit = True


def _set_width(ws, widths):
    """열 너비 조정

    Args:
        ws (obj): worksheet 오브젝트
        widths (dict or list): 너비 조정 설정값 ex) {'3': 10, '5': 14} / [10, 14, 12]
    Usages:
        widths = [10, 12, 8, 14]  # list 형태
        widths = {'1': 10, '2': 12, '3': 8, '5': 14}  # dict 형태, column number
        widths = {'A': 10, 'C': 12, 'D': 8, 'F': 14}  # dict 형태, column letter
        _set_width(ws, widths)
    """
    if type(widths) == dict:
        ## 'A', 'B' 형식인지, '1', '2', 형식인지 확인, isalpha(), isdigit()
        if list(widths.keys())[0].isalpha():
            for key, val in widths.items():
                ws.column_dimensions[key].width = val
        elif list(widths.keys())[0].isdigit():
            for key, val in widths.items():
                col = openpyxl.utils.get_column_letter(int(key) + 1)
                ws.column_dimensions[col].width = val
        else:
            pass

    elif type(widths) == list:
        column_letters = tuple(openpyxl.utils.get_column_letter(col_number + 1) for col_number in range(len(widths)))
        for i, column_letter in enumerate(column_letters):
            ws.column_dimensions[column_letter].width = widths[i]
    else:
        pass

def _set_height(ws, heights):
    """행 높이 조정

    Args:
        ws (obj): worksheet 오브젝트
        heights (dict or list): 너비 조정 설정값 ex) {'3': 10, '5': 14} / [10, 14, 12]
    Usages:
        heights = {1: 10, (2, 7): 20}  # dict 형태, int, tuple 사용
        _set_height(ws, heights)
    """
    for key, height in heights.items():
        if type(key) == int:
            ws.row_dimensions[key].height = height
        elif type(key) == tuple:
            for i in range(key[0], key[1] + 1):
#                 print(f"height {i}: {height}")
                ws.row_dimensions[i].height = height
        else:
            pass
            # print(f"oops!! key is ?: {key}")

            
def _put_rows(ws, rows, first=(0, 0)):
    """rows 값 입력

    Args:
        ws (obj): worksheet 오브젝트
        rows (list of list): row 리스트
        first (tuple, optional): 시작 셀 (행, 열). Defaults to (0, 0).
    Usages:
        rows = [
            [1, 2, 3, 4],
            [5, 6, 7, 8],
        ]
        _put_rows(ws, rows, first=(0, 0))
    """

    for i, row in enumerate(rows):
        for j, val in enumerate(row):
            ws.cell(row=i+first[0]+1, column=j+first[1]+1, value=val)


def _insert_image_cell(ws=None, path='', row=0, column=0, width=None, height=None):
    """이미지 삽입(셀 단위)

    Args:
        ws (obj, optional): [description]. Defaults to None.
        path (str, optional): 이미지 파일 경로. Defaults to ''.
        row (int, optional): 행 번호(시작번호: 0). Defaults to 0.
        column (int, optional): 열 번호(시작번호: 0) Defaults to 0.
        width (float, optional): 이미지 높이. Defaults to None.
        height (float, optional): 이미지 너비. Defaults to None.
    """
    img = openpyxl.drawing.image.Image(path)
    img.anchor = f"{openpyxl.utils.get_column_letter(column+1)}{row+1}"
    if width != None:
        img.width = width  # image 너비 조정
        # ws.column_dimensions[openpyxl.utils.get_column_letter(column+1)].width = width*WIDTH_PX_RATE  ## cell 너비 조정
    if height != None:
        img.height = height  # image 높이 조정
        # ws.row_dimensions[row+1].height = height*HEIGHT_PX_RATE  # cell 높이 조정

    ws.add_image(img)


def _insert_image_rows(ws=None, paths=[], rows=[0, 1], column=0, width=None, height=None):
    """이미지 삽입(행 단위, 열: 지정한 1열)

    Args:
        ws (obj, optional): [description]. Defaults to None.
        paths (list of str, optional): 이미지 파일 경로 리스트. Defaults to ''.
        rows (int, optional): 행 번호 리스트(시작번호: 0) 예) [0, 1, 2, 4]. Defaults to 0.
        column (int, optional): 열 번호(시작번호: 0) Defaults to 0.
        width (float, optional): 이미지 높이. Defaults to None.
        height (float, optional): 이미지 너비. Defaults to None.
    """
    images = zip(rows, paths)
    ws.column_dimensions[openpyxl.utils.get_column_letter(column+1)].width = width*WIDTH_PX_RATE  ## cell 너비 조정
    for row, path in images:
        _insert_image_cell(ws=ws, path=path, row=row, column=column, width=width, height=height)
        ws.row_dimensions[row+1].height = height*HEIGHT_PX_RATE  # cell 높이 조정


def _insert_image_cells(ws=None, paths=[[]], rows=[0, 1], columns=[], width=None, height=None):
    """이미지 삽입(행.렬 범위 지정)

    Args:
        ws (obj, optional): [description]. Defaults to None.
        path (str, optional): 이미지 파일 경로. Defaults to ''.
        rows (int, optional): 행 번호 리스트(시작번호: 0) 예) [0, 1, 2, 4]. Defaults to 0.
        columns (int, optional): 열 번호 리스트(시작번호: 0) 예) [0, 1, 2, 4] Defaults to 0.
        width (float, optional): 이미지 높이. Defaults to None.
        height (float, optional): 이미지 너비. Defaults to None.
    """
    images = zip(rows, paths)
    for row, _paths in images:
        for column, path in zip(columns, _paths):
            _insert_image_cell(ws=ws, path=path, row=row, column=column, width=width, height=height)


def _write_xsheet(path=None, sheet=None, data=[], theme=None, header=0, first=(0, 0), new=False, save=True):
#     print(f"{_fn(1)}")
    wb = sheet['wb']
    ws = sheet['ws']
    max_row = ws.max_row
    if max_row < 2 or new: ## 빈 sheet이거나, new(새로 만듬, 덮어씀)
        rows = _to_lists(data, header=header)
        # print(f"empty sheet!!: {max_row}")
        _put_rows(ws, rows, first=first)  ## 데이터 입력
        if theme != None: ## NOTE: theme 적용
            count_row = len(rows)
            count_col = len(rows[0])
            _apply_style_theme(sheet, theme, count_row, count_col, header, first)
    else:
        rows = _to_lists(data, header=None)
        # print(f"NOT empty sheet!!: {max_row}")
        _put_rows(ws, rows, first=(first[0]+max_row, first[1])) ## 빈 행부터 data 추가
        ## 기존 마지막 행의 스타일 적용
        if theme == None: ## NOTE: theme 적용
            count_row = len(rows)
            count_col = len(rows[0])
            _copy_before_row_style(ws, from_to=(first[0]+max_row, first[0]+max_row+count_row))
        else: ## NOTE: theme 적용
            count_row = len(rows) + max_row
            count_col = len(rows[0])
            _apply_style_theme(sheet, theme, count_row, count_col, header, first)

    _create_folder(path)  # 폴더 생성
    
    if save:
        if not wb.views:
            wb.views.append(openpyxl.workbook.views.BookView())
        wb.save(path)  # 저장
    else:
        return wb


def _read_xsheet(sheet=None, header=0, first=(0, 0), fields=None, project={}, out_type='dicts'):
    ws = sheet['ws']
    max_row = ws.max_row

    for i in range(max_row, 1, -1): ## NOTE: max_row부터 역으로 비어있지 않은 행을 찾음
        val = ws.cell(row=i, column=first[1]+1).value
        # print(f"val {i}: {val}")
        if val != None and str(val).strip() != '':
            # print(f"not empty: {i}")
            max_row = i
            break

    max_col = ws.max_column + 1  ## TODO: +1이 언제나 필요한지 확인 필요

    _header = [ws.cell(row=header+first[0]+1, column=column).value for column in range(first[1]+1, max_col)]
    if fields == None:  ## NOTE: 모든 fields 값들 출력
        _rows = [
            # ['' if ws.cell(row=row, column=column).value == None else  ws.cell(row=row, column=column).value
            [ws.cell(row=row, column=column).value
            for column in range(first[1]+1, max_col)]
            for row in range(header+first[0]+2, max_row+1)
        ]
        # print(f"_rows: {_rows}")
    else: ## NOTE: 해당 fields 값들만 출력
        # print(f"_header: {_header}")
        # print(f"fields: {fields}")
        exclusions = [i for i, _h in enumerate(_header) if not _h in fields]
        # print(f"exclusions: {exclusions}")
        _header = [ws.cell(row=header+first[0]+1, column=column).value for column in range(first[1]+1, max_col) if not column-1 in exclusions]
        _rows = [
            [ws.cell(row=row, column=column).value
                for column in range(first[1]+1, max_col)  if not column-1 in exclusions]
            for row in range(header+first[0]+2, max_row+1)
        ]
   
    data = [_header] + _rows

    if out_type == 'dicts':
        data = _to_dicts(data)
        # print(f"data: {data}")
        # print(f"project: {project}")
        if project != {} and project != None:
            # print(f"data: {data}")
            # print(f"project: {project}")
            data = _project_dicts(data, project)
    elif out_type == 'frame' or out_type == 'df':
        data = _to_df(data)

    return data


def write_xsheet(path='', title='', data=[], theme=None, header=0, first=(0, 0), new=False, save=True):
    sheet = _xsheet(path=path, title=title)
    _write_xsheet(path=path, sheet=sheet, data=data, theme=theme, header=header, first=first, new=new, save=save)


def read_xsheet(path='', title='', header=0, first=(0, 0), fields=None, project={}, out_type='dicts'):
    # print(f"path: {path}")
    # print(f"title: {title}")
    sheet = _xsheet(path=path, title=title)
    return _read_xsheet(sheet=sheet, header=header, first=first, fields=fields, project=project, out_type=out_type)
  

if __name__ == "__main__":
    ## NOTE: candle(wikipedia_org)
    xlsx_path = '../sats/_setup/_res/api_requests_ebest.xlsx'
    title = 'ebest_TR'

    # sheet = _xsheet(path=xlsx_path, title=title)
    # wb = sheet['wb']
    # ws = sheet['ws']

    fields = ['res_code', 'usage', 'priority', 'goods', 'remark']
    sh = read_xsheet(path=xlsx_path, title=title, header=0, first=(0, 0), fields=fields, out_type='dicts')
    # print(sh)

