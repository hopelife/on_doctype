import os, sys
from pprint import pprint
import json
import time
import re
import openpyxl
import pandas as pd

##------------------------------------------------------------
## User 모듈
##------------------------------------------------------------
sys.path.append(os.path.join(os.path.dirname(__file__), "."))
from .util_basic import (
    _create_folder, 
    _read_file, 
    _to_lists,
    _to_dicts,
    _to_df, 
    _filter_dicts,
    _fn
)

##@@@ 전역 상수/변수
##============================================================

##@@ setting, 
##------------------------------------------------------------

## openpyxl
##--------------------------------------------------------------
class Xlsx:
    def __init__(self, book_path: str=None, sheet_name: str=None) -> None:
        self.book_path = book_path
        # self.sheet_name = sheet_name
        self.set_book(book_path)
        self.set_sheet(sheet_name)
        self.set_range()  ## self.first(첫번째 비어있지 않은 셀), self.last(영역의 마지막 셀)
        self.sheetnames = self.book.sheetnames


    def set_book(self, book_path: str=None) -> None:
        """(work)book
        """
        if book_path == None:  # wb, ws 생성
            self.book = openpyxl.Workbook()
        else:  # wb, ws 로딩
            try:
                self.book = openpyxl.load_workbook(filename=book_path)
            except Exception as e:  ## NOTE: 파일이 없는 경우
                self.book = openpyxl.Workbook()


    def set_sheet(self, sheet_name: str=None) -> None:
        """(work)sheet
        """
        if sheet_name == None:  # sheet 이름이 없는 경우
            self.sheet = self.book.worksheets[0]
        try:
            if sheet_name in self.book.sheetnames:
                self.sheet = self.book[sheet_name]
            else:
                self.sheet = self.book.create_sheet(sheet_name) # sheet 생성
        except Exception as e:  ## NOTE: 에러시
            self.sheet = self.book.worksheets[0]

        self.sheet.sheet_name = sheet_name


    def set_range(self):
        rows = list(self.sheet.values)
        self.last = (len(rows), len(rows[0]))
        for i, row in enumerate(rows):
            for j, cell in enumerate(row):
                if cell != None:
                    self.first = (i, j)
                    return (self.first, self.last)


    def _put_rows(self, rows, first=(0, 0)):
        """rows 값 입력

        Args:
            ws (obj): worksheet 오브젝트
            rows (list of list): row 리스트
            first (tuple, optional): 시작 셀 (행, 열). Defaults to (0, 0).
        Usages:
            rows = [
                [1, 2, 3, 4],
                [5, 6, 7, 8],
            ]
            _put_rows(ws, rows, first=(0, 0))
        """
        for i, row in enumerate(rows):
            for j, val in enumerate(row):
                self.sheet.cell(row=i+first[0]+1, column=j+first[1]+1, value=val)


    def read(self, out_type='dicts'):
        data = list(self.sheet.values)
        # first ~ last data 추출
        data = [row[self.first[1]:] for row in data[self.first[0]:]]
        if out_type == "df":
            return _to_df(data)
        elif out_type == "dicts":
            return _to_dicts(data, header=0)
        else:
            return data


    def write(self, data, path=None, first=None):
        first = self.first if first == None else first
        path = self.book_path if path == None else path
        rows = _to_lists(data)
        self._put_rows(rows, first=self.first)
        self.book.save(filename = path)


if __name__ == "__main__":
    ## NOTE: load
    xlsx_path = "C:/Dev\docMoon/moonPackages/mp_util/_data/api_requests_ebest.xlsx"
    # xlsx_path = 'C:/Dev/docMoon/projects/sats/mp_ebest/api_manager/specs/api_requests_ebest.xlsx'
    sheet_name = 'desc'

    x = Xlsx(xlsx_path, sheet_name)

    ## NOTE: read
    # pprint(x.read(out_type="lists"))
    # print(f"first: {x.first}, last: {x.last}")
    
    ## NOTE: write
    data = [
        ('구분1', '위치1', '설명1', '비고1'),
        ('Res file', 'C:\\eBEST\\xingAPI\\Res', None, None),
        ('postgresql', 'api_requests_spec_ebest', 'spec 원본1', None),
        ('mongodb1', 'api_requests_user_ebest', '사용자 수정본', None),
        ('mariadb1', 'api_requests_spec_ebest_bk', 'spec 원본 백업', None)
    ]

    x.write(data)
